# -*- coding: utf-8 -*-
"""
超市采购对账系统 - Web 界面
启动: python web_app.py
"""

import os
import sys
import json
import time
import socket
import threading
import subprocess
import zipfile
from datetime import datetime
from pathlib import Path
from io import BytesIO

from flask import (Flask, render_template_string, request, jsonify,
                   Response, send_file, send_from_directory)
import openpyxl

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

# ─── 配置 ───
from config import BASE_DATA_DIR, BASE_APP_DIR, OUTPUT_DIR
SCRIPT_DIR = BASE_APP_DIR
BASE_DIR = ""  # 当前选中的批次路径，启动时由 _get_batches() 自动设置

# ─── 全局状态 ───
run_state = {
    "running": False,
    "logs": [],
    "current_supplier": "",
    "current_index": 0,
    "total": 0,
    "done": False,
    "results": [],
    "mode": "purchase",
}


def _get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def _get_suppliers():
    if not os.path.isdir(BASE_DIR):
        return []
    return sorted([d.name for d in Path(BASE_DIR).iterdir()
                   if d.is_dir() and not d.name.startswith("_")])


def _get_reports():
    if not os.path.isdir(OUTPUT_DIR):
        return []
    return sorted([f.name for f in Path(OUTPUT_DIR).iterdir()
                   if f.suffix == ".xlsx"])


# ═══════════════════════════════════════
# HTML 模板
# ═══════════════════════════════════════

STYLE = """
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, "Microsoft YaHei", sans-serif; background: #f5f5f5; color: #333; }
.container { max-width: 900px; margin: 0 auto; padding: 15px; }
h1 { text-align: center; padding: 20px 0 10px; color: #1a56db; font-size: 24px; }
.card { background: #fff; border-radius: 8px; padding: 20px; margin-bottom: 15px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
.card h2 { font-size: 16px; color: #555; margin-bottom: 12px; border-bottom: 1px solid #eee; padding-bottom: 8px; }
label { display: inline-block; margin: 4px 8px 4px 0; cursor: pointer; font-size: 14px; }
input[type=checkbox] { margin-right: 4px; }
.btn { display: inline-block; padding: 12px 36px; background: #1a56db; color: #fff; border: none;
       border-radius: 6px; font-size: 16px; cursor: pointer; text-decoration: none; }
.btn:hover { background: #1444b0; }
.btn:disabled, .btn.disabled { background: #999; cursor: not-allowed; }
.btn-sm { padding: 6px 16px; font-size: 13px; }
.btn-green { background: #16a34a; }
.btn-green:hover { background: #15803d; }
.center { text-align: center; }
.suppliers { display: flex; flex-wrap: wrap; gap: 0; }
.suppliers label { min-width: 100px; padding: 4px 0; }
.progress-bar { width: 100%; height: 24px; background: #e5e7eb; border-radius: 12px; overflow: hidden; margin: 10px 0; }
.progress-fill { height: 100%; background: #1a56db; transition: width 0.3s; border-radius: 12px;
                  display: flex; align-items: center; justify-content: center; color: #fff; font-size: 12px; }
.log-box { background: #1e1e1e; color: #d4d4d4; font-family: Consolas, monospace; font-size: 12px;
           padding: 12px; border-radius: 6px; height: 300px; overflow-y: auto; white-space: pre-wrap; word-break: break-all; }
table { width: 100%; border-collapse: collapse; font-size: 14px; }
th, td { padding: 8px 12px; text-align: left; border-bottom: 1px solid #eee; }
th { background: #f8f9fa; font-weight: 600; }
tr:hover { background: #f0f7ff; }
.tag { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 12px; }
.tag-green { background: #dcfce7; color: #166534; }
.tag-yellow { background: #fef9c3; color: #854d0e; }
.tag-red { background: #fee2e2; color: #991b1b; }
.btns-row { display: flex; gap: 8px; flex-wrap: wrap; margin: 8px 0; }
.status-text { font-size: 14px; color: #555; margin: 5px 0; }
@media (max-width: 600px) {
  .container { padding: 10px; }
  .suppliers label { min-width: 80px; font-size: 13px; }
  .btn { padding: 10px 24px; font-size: 14px; }
  .log-box { height: 200px; font-size: 11px; }
}
</style>
"""

INDEX_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>超市采购对账系统</title>""" + STYLE + """
<style>
.batch-select { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }
.batch-select select { padding: 8px 12px; border: 1px solid #d1d5db; border-radius: 6px; font-size: 14px; min-width: 320px; }
.btn-refresh { background: none; border: 1px solid #d1d5db; border-radius: 6px; padding: 7px 12px; cursor: pointer;
               font-size: 16px; color: #555; transition: all 0.2s; }
.btn-refresh:hover { border-color: #1a56db; color: #1a56db; background: #eff6ff; }
.btn-refresh.spinning { animation: spin 0.6s linear; pointer-events: none; }
@keyframes spin { from { transform: rotate(0deg); } to { transform: rotate(360deg); } }
</style>
</head><body>
<div class="container">
<h1>超市采购对账系统</h1>

<div class="card">
<form method="POST" action="/run" id="mainForm">
  <input type="hidden" name="batch_path" id="batchPathInput" value="{{current_batch}}">
  <h2 style="margin-bottom:12px">选择批次和供应商</h2>

  <div class="batch-select" style="margin-bottom:15px">
    <label>数据批次:</label>
    <select name="batch" id="batchSelect" onchange="loadSuppliers()">
      {% for b in batches %}
      <option value="{{b.path}}" {% if b.current %}selected{% endif %}>{{b.name}}</option>
      {% endfor %}
    </select>
    <button type="button" class="btn-refresh" onclick="refreshBatches()" id="refreshBtn" title="重新扫描目录">&#x1f504;</button>
  </div>

  <div id="supplierArea">
    <div class="btns-row" style="margin-bottom:8px">
      <button type="button" class="btn btn-sm" onclick="checkAll(true)">全选</button>
      <button type="button" class="btn btn-sm" onclick="checkAll(false)" style="background:#6b7280">取消全选</button>
    </div>
    <div class="suppliers">
      {% for s in suppliers %}
      <label><input type="checkbox" name="suppliers" value="{{s}}" checked> {{s}}</label>
      {% endfor %}
    </div>
  </div>

  <div style="margin-top:15px;padding:12px;background:#f8f9fa;border-radius:8px">
    <label style="font-weight:600;font-size:14px;display:block;margin-bottom:8px">对账模式：</label>
    <label style="display:inline-flex;align-items:center;gap:6px;margin-right:20px;cursor:pointer">
      <input type="radio" name="mode" value="purchase" checked style="accent-color:#1a56db">
      <span>采购对账（倩茹）</span>
      <span style="font-size:11px;color:#888">— 核对送货单 vs 采购入库单</span>
    </label>
    <br style="margin-bottom:6px">
    <label style="display:inline-flex;align-items:center;gap:6px;cursor:pointer;margin-top:6px">
      <input type="radio" name="mode" value="finance" style="accent-color:#1a56db">
      <span>财务对账（文钰）</span>
      <span style="font-size:11px;color:#888">— 核对送货单 vs 进货单</span>
    </label>
  </div>

  <div style="margin-top:12px">
    <label><input type="checkbox" name="no_cache"> 强制重新OCR（清除缓存）</label>
  </div>

  <div class="center" style="margin-top:15px">
    <button type="submit" class="btn">开始对账</button>
  </div>
  <div style="margin-top:12px;text-align:center;font-size:12px;color:#aaa">
    本机局域网地址：<a href="http://{{local_ip}}:5000" style="color:#888">http://{{local_ip}}:5000</a>
    &nbsp;（文钰可在自己电脑浏览器输入此地址访问）
  </div>
</form>
</div>

{% if has_reports %}
<div class="card">
  <h2>历史报告</h2>
  <a href="/result" class="btn btn-sm btn-green">查看最新结果</a>
</div>
{% endif %}

</div>
<script>
function checkAll(v) {
  document.querySelectorAll('input[name=suppliers]').forEach(function(c) { c.checked = v; });
}

function loadSuppliers() {
  var sel = document.getElementById('batchSelect');
  var batchPath = sel.value;
  fetch('/api/suppliers?batch=' + encodeURIComponent(batchPath))
    .then(function(r) { return r.json(); })
    .then(function(data) {
      var area = document.querySelector('#supplierArea .suppliers');
      var html = '';
      data.suppliers.forEach(function(s) {
        html += '<label><input type="checkbox" name="suppliers" value="' + s + '" checked> ' + s + '</label>';
      });
      area.innerHTML = html || '<span style="color:#888">该批次下无供应商文件夹</span>';
      document.getElementById('batchPathInput').value = batchPath;
    });
}

function refreshBatches() {
  var btn = document.getElementById('refreshBtn');
  btn.classList.add('spinning');
  fetch('/api/batches')
    .then(function(r) { return r.json(); })
    .then(function(data) {
      var sel = document.getElementById('batchSelect');
      var curVal = sel.value;
      sel.innerHTML = '';
      data.batches.forEach(function(b) {
        var opt = document.createElement('option');
        opt.value = b.path;
        opt.textContent = b.name;
        if (b.path === curVal) opt.selected = true;
        sel.appendChild(opt);
      });
      // 如果之前选中的批次不在新列表中，选第一个
      if (sel.selectedIndex < 0 && sel.options.length > 0) {
        sel.selectedIndex = 0;
      }
      loadSuppliers();
      btn.classList.remove('spinning');
    })
    .catch(function() {
      btn.classList.remove('spinning');
      alert('刷新失败');
    });
}
</script>
</body></html>
"""

RUN_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>对账进行中</title>""" + STYLE + """</head><body>
<div class="container">
<h1>对账进行中</h1>

<div class="card">
  <p class="status-text" id="status">准备中...</p>
  <div class="progress-bar"><div class="progress-fill" id="pbar" style="width:0%"></div></div>
</div>

<div class="card">
  <h2>处理日志</h2>
  <div class="log-box" id="logbox"></div>
</div>

<div class="center" id="done-area" style="display:none">
  <a href="/result" class="btn btn-green">查看对账结果</a>
  <a href="/" class="btn" style="background:#6b7280;margin-left:10px">返回首页</a>
</div>
</div>
<script>
var logLen = 0;
function poll() {
  fetch('/api/status').then(r=>r.json()).then(data => {
    document.getElementById('status').textContent = data.status;
    var pct = data.total > 0 ? Math.round(data.current / data.total * 100) : 0;
    var pbar = document.getElementById('pbar');
    pbar.style.width = pct + '%';
    pbar.textContent = pct + '%';

    if (data.logs.length > logLen) {
      var box = document.getElementById('logbox');
      for (var i = logLen; i < data.logs.length; i++) {
        box.textContent += data.logs[i] + '\\n';
      }
      logLen = data.logs.length;
      box.scrollTop = box.scrollHeight;
    }

    if (data.done) {
      document.getElementById('done-area').style.display = 'block';
      document.getElementById('status').textContent = '对账完成！';
    } else {
      setTimeout(poll, 1500);
    }
  });
}
poll();
</script>
</body></html>
"""

RESULT_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>对账结果</title>""" + STYLE + """<style>
.overview { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 10px; }
.stat-box { text-align: center; padding: 14px 8px; border-radius: 8px; background: #f8f9fa; }
.stat-box .num { font-size: 26px; font-weight: bold; }
.stat-box .label { font-size: 11px; color: #666; margin-top: 4px; }
.stat-box.red { background: #fee2e2; } .stat-box.red .num { color: #dc2626; }
.stat-box.green .num { color: #16a34a; } .stat-box.blue .num { color: #1a56db; }
.todo-list { max-height: 320px; overflow-y: auto; }
.todo-row { display: flex; align-items: center; gap: 8px; padding: 6px 4px; border-bottom: 1px solid #f0f0f0; font-size: 13px; }
.todo-row:hover { background: #fef9c3; }
.todo-row.done label { text-decoration: line-through; color: #aaa; }
.todo-row .todo-supplier { min-width: 50px; font-weight: 600; color: #1a56db; }
.todo-row .todo-bc { min-width: 110px; font-family: Consolas, monospace; font-size: 12px; color: #555; }
.todo-row .todo-name { flex: 1; min-width: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.todo-row .todo-desc { color: #dc2626; font-weight: 500; min-width: 200px; }
.s-card { background: #fff; border-radius: 8px; padding: 16px; margin-bottom: 12px;
          box-shadow: 0 1px 3px rgba(0,0,0,0.1); border-left: 4px solid #ccc; }
.s-card.green { border-left-color: #16a34a; } .s-card.yellow { border-left-color: #eab308; } .s-card.red { border-left-color: #dc2626; }
.s-top { display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 8px; }
.s-name { font-size: 17px; font-weight: bold; }
.s-status { font-size: 13px; }
.s-body { display: flex; flex-wrap: wrap; gap: 15px; align-items: center; margin: 8px 0; }
.s-rate { text-align: center; }
.s-rate .val { font-size: 22px; font-weight: bold; } .s-rate .lbl { font-size: 11px; color: #888; }
.s-info { font-size: 13px; color: #555; line-height: 1.8; }
.s-info span { font-weight: 600; }
.s-diff-badge { display: inline-block; padding: 2px 10px; border-radius: 12px; font-size: 13px; font-weight: 600; cursor: pointer; }
.s-diff-badge.diff-green { background: #dcfce7; color: #166534; }
.s-diff-badge.diff-orange { background: #ffedd5; color: #c2410c; }
.s-diff-badge.diff-orange-red { background: #fee2e2; color: #dc2626; }
.s-diff-badge.diff-red { background: #dc2626; color: #fff; }
.s-actions { display: flex; gap: 6px; flex-wrap: wrap; }
.diff-detail { display: none; margin-top: 10px; border-top: 1px solid #eee; padding-top: 10px; }
.diff-detail.show { display: block; }
.diff-table { width: 100%; font-size: 12px; border-collapse: collapse; }
.diff-table th { background: #f8f9fa; padding: 6px 8px; text-align: left; font-size: 11px; }
.diff-table td { padding: 5px 8px; border-bottom: 1px solid #f0f0f0; }
.diff-table tr:hover { background: #fff7ed; }
.diff-table .sev-red { color: #dc2626; font-weight: 600; }
.diff-table .sev-yellow { color: #b45309; font-weight: 600; }
.empty-msg { text-align: center; padding: 60px 20px; color: #999; font-size: 16px; }
.tip-icon { display: inline-flex; align-items: center; justify-content: center; width: 16px; height: 16px;
            border-radius: 50%; background: #9ca3af; color: #fff; font-size: 11px; font-weight: bold;
            cursor: help; margin-left: 4px; vertical-align: middle; position: relative; }
.tip-icon:hover::after { content: attr(title); position: absolute; top: 22px; left: 50%; transform: translateX(-50%);
            background: #333; color: #fff; padding: 6px 10px; border-radius: 6px; font-size: 12px; font-weight: normal;
            white-space: nowrap; z-index: 100; max-width: 280px; white-space: normal; line-height: 1.4;
            box-shadow: 0 2px 8px rgba(0,0,0,0.2); }
@media (max-width: 600px) {
  .s-body { gap: 10px; }
  .todo-row .todo-desc { min-width: 120px; font-size: 12px; }
  .diff-table { font-size: 11px; }
}
</style></head><body>
<div class="container">
<h1>{% if mode == 'finance' %}财务对账结果（文钰）{% else %}采购对账结果（倩茹）{% endif %}</h1>

{% if not data %}
<div class="card"><div class="empty-msg">暂无数据，请先<a href="/">开始对账</a></div></div>
{% else %}

<!-- 概览 -->
<div class="card">
  <h2>本次对账概览</h2>
  <div class="overview">
    <div class="stat-box"><div class="num">{{summary.date}}</div><div class="label">对账日期</div></div>
    <div class="stat-box blue"><div class="num">{{summary.total_suppliers}}</div><div class="label">供应商数</div></div>
    {% if mode != 'finance' %}
    <div class="stat-box green"><div class="num">{{summary.purchase_rate}}</div><div class="label">采购一致率</div></div>
    {% endif %}
    {% if mode != 'purchase' %}
    <div class="stat-box green"><div class="num">{{summary.finance_rate}}</div><div class="label">财务一致率</div></div>
    {% endif %}
    <div class="stat-box {% if summary.total_diff > 0 %}red{% endif %}">
      <div class="num">{{summary.total_diff}}</div><div class="label">差异待处理</div></div>
    {% if mode != 'purchase' %}
    <div class="stat-box blue"><div class="num">{{summary.total_payable}}</div><div class="label">合计应付</div></div>
    {% if summary.total_return != '0.00' %}<div class="stat-box"><div class="num">{{summary.total_return}}</div><div class="label">合计退货</div></div>{% endif %}
    {% endif %}
  </div>
  {% if po_numbers %}
  <div style="margin-top:12px;font-size:13px;color:#555;line-height:1.8">
    <span style="font-weight:600">本批次PO号：</span>
    <span style="color:#333">{{po_numbers|join('、')}}</span>
  </div>
  {% endif %}
</div>

<!-- 待处理事项 -->
{% if todo_groups %}
<div class="card">
  <h2 style="color:#dc2626">待处理事项 ({{todo_total}} 条差异)</h2>
  <div class="btns-row" style="margin-bottom:10px">
    <button class="btn btn-sm" style="background:#dc2626" onclick="filterTodo('red')" id="fbtn-red">仅严重差异</button>
    <button class="btn btn-sm" style="background:#6b7280" onclick="filterTodo('all')" id="fbtn-all">全部</button>
    <span style="font-size:12px;color:#888;line-height:32px;margin-left:8px">采购助理工作清单 — 勾选表示已处理</span>
  </div>
  {% for g in todo_groups %}
  <div class="todo-group" data-supplier="{{g.supplier}}">
    <div class="todo-group-header" onclick="toggleGroup(this)" style="cursor:pointer;padding:8px 4px;border-bottom:1px solid #eee;display:flex;align-items:center;gap:8px">
      <span class="todo-arrow" style="font-size:12px;color:#888">&#9654;</span>
      <span style="font-weight:600;color:#1a56db">{{g.supplier}}</span>
      <span class="s-diff-badge {% if g.entries|length >= 50 %}diff-red{% elif g.entries|length >= 10 %}diff-orange-red{% else %}diff-orange{% endif %}" style="font-size:12px">{{g.entries|length}} 条差异</span>
    </div>
    <div class="todo-group-body" style="display:none">
      {% for t in g.entries %}
      <div class="todo-row {% if not loop.first and loop.index > 5 %}todo-extra{% endif %}" data-sev="{{t.sev}}" {% if loop.index > 5 %}style="display:none"{% endif %}>
        <input type="checkbox" onchange="this.parentElement.classList.toggle('done')">
        <span class="todo-bc">{{t.barcode}}</span>
        <span class="todo-name">{{t.name}}</span>
        <span class="todo-desc">{{t.desc}}</span>
      </div>
      {% endfor %}
      {% if g.entries|length > 5 %}
      <div class="todo-show-all" style="text-align:center;padding:6px">
        <button class="btn btn-sm" style="background:#6b7280;font-size:12px" onclick="showAllInGroup(this, {{g.entries|length}})">显示全部 {{g.entries|length}} 条</button>
      </div>
      {% endif %}
    </div>
  </div>
  {% endfor %}
</div>
{% endif %}

<!-- 供应商卡片 -->
<div class="card">
  <h2>各供应商明细</h2>
  {% for s in data %}
  <div class="s-card {{s.color}}">
    <div class="s-top">
      <span class="s-name">{{s.supplier}}</span>
      <span class="s-status">{{s.status|safe}}</span>
    </div>
    <div class="s-body">
      {% if mode != 'finance' %}
      <div class="s-rate"><div class="val" style="color:{{s.purchase_color}}">{{s.purchase_rate}}</div><div class="lbl">采购对账</div></div>
      {% endif %}
      {% if mode != 'purchase' %}
      <div class="s-rate"><div class="val" style="color:{{s.finance_color}}">{{s.finance_rate}}</div><div class="lbl">财务对账</div></div>
      <div class="s-info">
        进货: <span>{{s.goods_amount}}</span>
        {% if s.return_amount != '0.00' %} &nbsp; 退货：<span style="color:#dc2626;font-weight:600">-&yen;{{s.return_amount}}</span>{% endif %}
        &nbsp; 应付: <span style="color:#1a56db">{{s.payable}}</span>
      </div>
      {% endif %}
    </div>
    <div class="s-actions">
      {% if s.diff_count == 0 %}
        <span class="s-diff-badge diff-green">&#10003; 无差异</span>
      {% elif s.diff_count < 10 %}
        <span class="s-diff-badge diff-orange" onclick="toggleDiff('diff-{{loop.index}}')">{{s.diff_count}} 条差异 &#9660;</span>
      {% elif s.diff_count < 50 %}
        <span class="s-diff-badge diff-orange-red" onclick="toggleDiff('diff-{{loop.index}}')">{{s.diff_count}} 条差异 &#9660;</span>
      {% else %}
        <span class="s-diff-badge diff-red" onclick="toggleDiff('diff-{{loop.index}}')">{{s.diff_count}} 条差异 &#9660;</span>
      {% endif %}
      {% if s.report_file %}<a href="/download/{{s.report_file}}" class="btn btn-sm">下载报告</a>{% endif %}
    </div>
    {% if s.diffs %}
    <div class="diff-detail" id="diff-{{loop.index}}">
      <table class="diff-table">
        <tr><th>条形码</th><th>商品名称</th><th>差异类型</th><th>送货单</th><th>入库单</th><th>差值</th><th>级别</th></tr>
        {% for d in s.diffs %}
        <tr>
          <td style="font-family:Consolas;font-size:11px">{{d.barcode}}</td>
          <td>{{d.name}}</td>
          <td>{{d.diff_type}}</td>
          <td>{{d.delivery_val}}</td>
          <td>{{d.receipt_val}}</td>
          <td>{{d.diff_val}}</td>
          <td class="sev-{{d.sev_class}}">{{d.severity}}</td>
        </tr>
        {% endfor %}
      </table>
    </div>
    {% endif %}
  </div>
  {% endfor %}
</div>

<!-- 快捷操作 -->
<div class="card">
  <h2>快捷操作</h2>
  <div class="btns-row">
    <a href="/download_all" class="btn btn-sm btn-green">全部报告打包下载 (ZIP)</a>
    <a href="/download_payable_summary" class="btn btn-sm">导出应付款汇总表</a>
  </div>
</div>

{% endif %}
<div class="center" style="margin-top:15px"><a href="/" class="btn" style="background:#6b7280">返回首页</a></div>
</div>
<script>
function toggleDiff(id) {
  var el = document.getElementById(id);
  el.classList.toggle('show');
}
function toggleGroup(header) {
  var body = header.nextElementSibling;
  var arrow = header.querySelector('.todo-arrow');
  if (body.style.display === 'none') {
    body.style.display = 'block';
    arrow.innerHTML = '&#9660;';
  } else {
    body.style.display = 'none';
    arrow.innerHTML = '&#9654;';
  }
}
function showAllInGroup(btn, total) {
  var body = btn.closest('.todo-group-body');
  body.querySelectorAll('.todo-extra').forEach(function(el) { el.style.display = ''; });
  btn.parentElement.style.display = 'none';
}
function filterTodo(mode) {
  document.querySelectorAll('.todo-row').forEach(function(el) {
    if (mode === 'red') {
      el.style.display = el.dataset.sev === 'red' ? '' : 'none';
    } else {
      // 还原：前5条显示，extra隐藏（除非已展开）
      el.style.display = '';
    }
  });
  document.getElementById('fbtn-red').style.background = mode === 'red' ? '#dc2626' : '#6b7280';
  document.getElementById('fbtn-all').style.background = mode === 'all' ? '#1a56db' : '#6b7280';
}
</script>
</body></html>
"""


# ═══════════════════════════════════════
# 路由
# ═══════════════════════════════════════

def _is_batch_dir(d):
    """判断目录是否为有效批次（含供应商子文件夹且子文件夹内有Excel/PDF）"""
    if not d.is_dir():
        return False
    subs = [s for s in d.iterdir() if s.is_dir() and not s.name.startswith(("_", "."))]
    if not subs:
        return False
    return any((s / f).exists() for s in subs for f in os.listdir(s) if f.endswith(('.xlsx', '.pdf')))

def _is_supplier_dir(d):
    """判断目录是否为供应商目录（直接包含 xlsx/pdf 文件）"""
    if not d.is_dir():
        return False
    return any(f.suffix.lower() in ('.xlsx', '.xls', '.pdf') for f in d.iterdir() if f.is_file())

def _get_supplier_names(batch_path):
    """获取批次下的供应商名称列表（排除嵌套的子批次）"""
    p = Path(batch_path)
    names = []
    for s in p.iterdir():
        if not s.is_dir() or s.name.startswith(("_", ".")):
            continue
        # 供应商目录 = 直接包含 xlsx/pdf 文件
        if _is_supplier_dir(s):
            names.append(s.name)
    return sorted(names)

def _format_batch_name(dir_name):
    """尝试把目录名格式化为可读日期，如 3.16-对账汇总 → 2026-03-16"""
    import re
    year = datetime.now().year
    # 匹配 YYYYMMDD
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', dir_name)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # 匹配 M.DD 或 MM.DD 格式（3.16 → 月=3,日=16）
    m = re.match(r'^(\d{1,2})\.(\d{1,2})', dir_name)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        return f"{year}-{month:02d}-{day:02d}"
    return dir_name

def _get_batches():
    """扫描 BASE_DATA_DIR 下所有有效批次（两层深度），按名称倒序"""
    batches = []
    scan_root = Path(BASE_DATA_DIR)
    if not scan_root.is_dir():
        return batches

    for d in sorted(scan_root.iterdir(), reverse=True):
        if not d.is_dir() or d.name.startswith(("_", ".", "__")):
            continue
        if _is_batch_dir(d):
            suppliers = _get_supplier_names(str(d))
            display = _format_batch_name(d.name)
            if suppliers:
                display += f"（{'、'.join(suppliers[:5])}{'…' if len(suppliers) > 5 else ''}）"
            batches.append({"name": display, "path": str(d), "count": len(suppliers),
                            "suppliers": suppliers, "current": str(d) == BASE_DIR})
            # 扫描子目录一层
            for sub in sorted(d.iterdir(), reverse=True):
                if sub.is_dir() and not sub.name.startswith(("_", ".")) and _is_batch_dir(sub):
                    sub_suppliers = _get_supplier_names(str(sub))
                    sub_display = _format_batch_name(sub.name)
                    if sub_suppliers:
                        sub_display += f"（{'、'.join(sub_suppliers[:5])}{'…' if len(sub_suppliers) > 5 else ''}）"
                    batches.append({"name": sub_display, "path": str(sub), "count": len(sub_suppliers),
                                    "suppliers": sub_suppliers, "current": str(sub) == BASE_DIR})
    return batches


@app.route("/")
def index():
    global BASE_DIR
    batches = _get_batches()

    # 支持批次切换
    batch_path = request.args.get("batch", BASE_DIR)
    if batch_path and os.path.isdir(batch_path):
        BASE_DIR = batch_path
    # 如果 BASE_DIR 未设置或无效，自动选第一个批次
    if (not BASE_DIR or not os.path.isdir(BASE_DIR)) and batches:
        BASE_DIR = batches[0]["path"]

    suppliers = _get_suppliers()
    has_reports = len(_get_reports()) > 0
    # 标记当前批次
    for b in batches:
        b["current"] = b["path"] == BASE_DIR
    return render_template_string(INDEX_HTML, suppliers=suppliers, has_reports=has_reports,
                                   batches=batches, current_batch=BASE_DIR,
                                   local_ip=_get_local_ip())


@app.route("/api/suppliers")
def api_suppliers():
    batch_path = request.args.get("batch", BASE_DIR)
    if not os.path.isdir(batch_path):
        return jsonify({"suppliers": []})
    subs = sorted([d.name for d in Path(batch_path).iterdir()
                   if d.is_dir() and not d.name.startswith(("_", "."))])
    return jsonify({"suppliers": subs})


@app.route("/api/batches")
def api_batches():
    batches = _get_batches()
    return jsonify({"batches": [{"name": b["name"], "path": b["path"]} for b in batches]})


@app.route("/api/create_batch", methods=["POST"])
def api_create_batch():
    data = request.get_json()
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"success": False, "msg": "请输入批次名称"})
    batch_path = os.path.join(SCRIPT_DIR, name)
    if os.path.exists(batch_path):
        return jsonify({"success": True, "path": batch_path, "msg": "批次已存在，继续使用"})
    os.makedirs(batch_path, exist_ok=True)
    return jsonify({"success": True, "path": batch_path})


@app.route("/api/create_supplier", methods=["POST"])
def api_create_supplier():
    data = request.get_json()
    batch = data.get("batch", "")
    supplier = data.get("supplier", "").strip()
    if not batch or not supplier:
        return jsonify({"success": False, "msg": "参数缺失"})
    supplier_path = os.path.join(batch, supplier)
    os.makedirs(supplier_path, exist_ok=True)
    return jsonify({"success": True})


@app.route("/api/upload", methods=["POST"])
def api_upload():
    batch = request.form.get("batch", "")
    supplier = request.form.get("supplier", "")
    file_type = request.form.get("type", "")
    files = request.files.getlist("files")

    if not batch or not supplier or not files:
        return jsonify({"success": False, "msg": "参数缺失"})

    supplier_path = os.path.join(batch, supplier)
    os.makedirs(supplier_path, exist_ok=True)

    saved = []
    for f in files:
        if not f.filename:
            continue
        # 保存文件
        filepath = os.path.join(supplier_path, f.filename)
        f.save(filepath)
        saved.append(f.filename)

    if not saved:
        return jsonify({"success": False, "msg": "没有有效文件"})

    return jsonify({"success": True, "files": saved})


@app.route("/run", methods=["GET", "POST"])
def run():
    global BASE_DIR
    if request.method == "POST" and not run_state["running"]:
        batch_path = request.form.get("batch_path", BASE_DIR)
        if os.path.isdir(batch_path):
            BASE_DIR = batch_path
        selected = request.form.getlist("suppliers")
        no_cache = "no_cache" in request.form
        mode = request.form.get("mode", "purchase")
        if selected:
            _start_reconcile(selected, no_cache, mode)
    return render_template_string(RUN_HTML)


@app.route("/api/status")
def api_status():
    s = run_state
    if s["running"]:
        status = f"正在处理: {s['current_supplier']} ({s['current_index']}/{s['total']})"
    elif s["done"]:
        status = "对账完成"
    else:
        status = "就绪"
    return jsonify({
        "status": status,
        "current": s["current_index"],
        "total": s["total"],
        "done": s["done"],
        "running": s["running"],
        "logs": s["logs"],
    })


@app.route("/result")
def result():
    mode = run_state.get("mode", "purchase")
    report_data = _load_summary_from_xlsx()
    if not report_data:
        return render_template_string(RESULT_HTML, data=None, summary=None, todo_items=None, mode=mode)

    data, summary = report_data

    # 读取每家供应商的差异明细，构建分组
    todo_groups = []
    todo_total = 0
    for s in data:
        diffs = _load_diffs_for_supplier(s["supplier"])
        s["diffs"] = diffs

        group_items = []
        for d in diffs:
            desc = _make_human_desc(d)
            group_items.append({
                "barcode": d["barcode"],
                "name": d["name"],
                "desc": desc,
                "sev": d["sev_class"],
            })

        if group_items:
            todo_groups.append({
                "supplier": s["supplier"],
                "entries": group_items,
            })
            todo_total += len(group_items)

    report_files = [s["report_file"] for s in data if s.get("report_file")]
    po_numbers = _collect_po_numbers(report_files)
    return render_template_string(RESULT_HTML, data=data, summary=summary,
                                  todo_groups=todo_groups, todo_total=todo_total, mode=mode,
                                  po_numbers=po_numbers)


@app.route("/download/<filename>")
def download(filename):
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)


@app.route("/download_all")
def download_all():
    reports = _get_reports()
    if not reports:
        return "没有报告文件", 404
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in reports:
            zf.write(os.path.join(OUTPUT_DIR, f), f)
    buf.seek(0)
    date_str = datetime.now().strftime("%Y%m%d")
    return send_file(buf, mimetype="application/zip",
                     as_attachment=True,
                     download_name=f"对账报告_{date_str}.zip")


@app.route("/download_payable_summary")
def download_payable_summary():
    """导出应付款汇总表"""
    report_data = _load_summary_from_xlsx()
    if not report_data:
        return "暂无数据", 404

    data, summary = report_data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "应付款汇总"
    headers = ["供应商", "进货金额", "退货金额", "应付净额", "采购对账一致率", "状态"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h).font = openpyxl.styles.Font(bold=True)
    for r, s in enumerate(data, 2):
        ws.cell(row=r, column=1, value=s["supplier"])
        ws.cell(row=r, column=2, value=float(s["goods_amount"].replace(",", "")))
        ws.cell(row=r, column=3, value=float(s["return_amount"].replace(",", "")))
        ws.cell(row=r, column=4, value=float(s["payable"].replace(",", "")))
        ws.cell(row=r, column=5, value=s["purchase_rate"])
        ws.cell(row=r, column=6, value=s["status_text"])
    # 合计行
    total_row = len(data) + 2
    ws.cell(row=total_row, column=1, value="合计").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=total_row, column=4, value=float(summary["total_payable"].replace(",", "")))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = summary.get("date", datetime.now().strftime("%Y%m%d"))
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"应付款汇总_{date_str}.xlsx")


def _make_human_desc(d):
    """将差异数据转为口语化描述"""
    diff_type = d.get("diff_type", "")
    dv = d.get("delivery_val", "")
    rv = d.get("receipt_val", "")
    diff_val = d.get("diff_val", "")

    parts = []
    if "数量" in diff_type and "单价" in diff_type:
        parts.append(f"数量和单价均不符：送货单{dv}，系统{rv}")
    elif "数量" in diff_type:
        parts.append(f"数量不符：送货单{dv}，系统{rv}，差{diff_val}")
    elif "单价" in diff_type:
        parts.append(f"单价不符：送货单{dv}元，系统{rv}元，差{diff_val}元")
    elif "金额" in diff_type:
        parts.append(f"金额不符：送货单{dv}元，系统{rv}元，差{diff_val}元")

    if not parts:
        if dv and rv:
            parts.append(f"数值差异：送货单{dv}，系统{rv}")
        else:
            parts.append("存在差异需核实")

    return "；".join(parts)


def _collect_po_numbers(report_files):
    """从指定的供应商报告文件中收集去重的PO号列表"""
    if not report_files:
        return []
    po_set = set()
    for fname in report_files:
        f = Path(OUTPUT_DIR) / fname
        if not f.exists():
            continue
        try:
            wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
            for sheet_name in ["一致项", "差异项"]:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                # 一致项: PO号在col 2, 差异项: PO号在col 3
                po_col = 2 if sheet_name == "一致项" else 3
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > po_col and row[po_col]:
                        po = str(row[po_col]).strip()
                        if po and po != "None":
                            po_set.add(po)
            wb.close()
        except Exception:
            continue
    return sorted(po_set)


def _load_diffs_for_supplier(supplier_name):
    """从供应商对账报告的Sheet2(差异项)读取差异明细"""
    if not os.path.isdir(OUTPUT_DIR):
        return []
    # 找报告文件
    report_files = [f for f in Path(OUTPUT_DIR).iterdir()
                    if supplier_name in f.name and "汇总" not in f.name and f.suffix == ".xlsx"]
    if not report_files:
        return []

    diffs = []
    try:
        wb = openpyxl.load_workbook(str(report_files[0]), read_only=True, data_only=True)
        if "差异项" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["差异项"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return []

        # 表头: 序号(0) 对账阶段(1) 严重级别(2) PO号(3) 存货编码(4) 商品名称我方(5)
        #        商品名称送货单(6) 仓库(7) 差异字段(8) 送货单值(9) 我方值(10) 差值(11) ...
        for row in rows[1:]:
            if not row[4]:
                continue
            severity = str(row[2]) if row[2] else ""
            sev_class = "red" if "红" in severity else "yellow"

            barcode = str(row[4]).strip() if row[4] else ""
            name = str(row[5]).strip() if row[5] else ""
            if not name and row[6]:
                name = str(row[6]).strip()
            # 截断过长名称
            if len(name) > 25:
                name = name[:25] + "..."

            diff_type = str(row[8]).strip() if row[8] else ""
            delivery_val = str(row[9]) if row[9] is not None else ""
            receipt_val = str(row[10]) if row[10] is not None else ""
            diff_val = str(row[11]) if row[11] is not None else ""

            diffs.append({
                "barcode": barcode,
                "name": name,
                "diff_type": diff_type,
                "delivery_val": delivery_val,
                "receipt_val": receipt_val,
                "diff_val": diff_val,
                "severity": severity,
                "sev_class": sev_class,
            })
    except Exception:
        pass

    return diffs


def _load_summary_from_xlsx():
    """从汇总报告xlsx读取数据"""
    # 找最新的汇总报告
    if not os.path.isdir(OUTPUT_DIR):
        return None
    summary_files = sorted([f for f in Path(OUTPUT_DIR).iterdir()
                           if "汇总" in f.name and f.suffix == ".xlsx"], reverse=True)
    if not summary_files:
        return None

    summary_file = summary_files[0]
    # 提取日期后缀用于匹配同批次的供应商报告（如 20260323）
    import re as _re
    _date_match = _re.search(r'(\d{8})', summary_file.name)
    report_date = _date_match.group(1) if _date_match else ""

    wb = openpyxl.load_workbook(str(summary_file), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return None

    header = rows[0]
    data_rows = rows[1:]
    reports = _get_reports()

    suppliers = []
    total_diff = 0
    total_payable = 0
    total_return = 0
    total_suppliers = 0
    date_str = ""

    for row in data_rows:
        name = str(row[0]) if row[0] else ""
        if name == "合计" or not name:
            continue

        total_suppliers += 1
        date_str = str(row[1]) if row[1] else ""

        purchase_rate = str(row[9]) if row[9] else "0%"
        finance_rate = str(row[13]) if row[13] else "0%"
        purchase_diff = int(row[7]) if row[7] else 0
        finance_diff = int(row[11]) if row[11] else 0
        diff_count = purchase_diff + finance_diff
        total_diff += diff_count

        goods_amt = float(row[14]) if row[14] else 0
        return_amt = abs(float(row[16])) if row[16] else 0
        payable = float(row[17]) if row[17] else 0
        total_payable += payable
        total_return += return_amt

        pct = _pct(purchase_rate)
        if diff_count == 0:
            color, status, status_text = "green", "✅ 无需处理", "正常"
        elif pct >= 85:
            color, status, status_text = "yellow", '⚠️ 有差异项，请处理后提交 <span class="tip-icon" title="该供应商存在数量/金额差异条目，需核对后在系统中修正再提交">?</span>', "待确认"
        else:
            color, status, status_text = "red", '❌ 一致率偏低，请重点核查 <span class="tip-icon" title="该供应商采购对账一致率低于85%，建议逐条核对差异项和未匹配项">?</span>', "需介入"

        def _rate_color(rate_str):
            p = _pct(rate_str)
            if p >= 90: return "#16a34a"
            if p >= 70: return "#eab308"
            return "#dc2626"

        # 找对应的报告文件（同日期+同供应商名）
        report_file = ""
        for f in reports:
            if name in f and "汇总" not in f and report_date in f:
                report_file = f
                break

        suppliers.append({
            "supplier": name,
            "purchase_rate": purchase_rate,
            "finance_rate": finance_rate,
            "purchase_color": _rate_color(purchase_rate),
            "finance_color": _rate_color(finance_rate),
            "diff_count": diff_count,
            "goods_amount": f"{goods_amt:,.2f}",
            "return_amount": f"{return_amt:,.2f}",
            "payable": f"{payable:,.2f}",
            "color": color,
            "status": status,
            "status_text": status_text,
            "report_file": report_file,
        })

    # 格式化日期
    if len(date_str) == 8:
        display_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
    else:
        display_date = date_str

    # 合计行数据
    totals_row = [r for r in data_rows if r[0] == "合计"]
    overall_purchase_rate = str(totals_row[0][9]) if totals_row else "-"
    overall_finance_rate = str(totals_row[0][13]) if totals_row else "-"

    summary = {
        "date": display_date,
        "total_suppliers": total_suppliers,
        "purchase_rate": overall_purchase_rate,
        "finance_rate": overall_finance_rate,
        "total_diff": total_diff,
        "total_payable": f"{total_payable:,.2f}",
        "total_return": f"{total_return:,.2f}",
    }

    return suppliers, summary


# ═══════════════════════════════════════
# 后台对账
# ═══════════════════════════════════════

def _start_reconcile(selected_suppliers, no_cache, mode="purchase"):
    run_state["running"] = True
    run_state["done"] = False
    run_state["logs"] = []
    run_state["current_supplier"] = ""
    run_state["current_index"] = 0
    run_state["total"] = len(selected_suppliers)
    run_state["results"] = []
    run_state["mode"] = mode

    thread = threading.Thread(target=_run_process, args=(selected_suppliers, no_cache, mode), daemon=True)
    thread.start()


def _run_process(selected_suppliers, no_cache, mode="purchase"):
    try:
        cmd = [sys.executable, "-X", "utf8", "-u",
               os.path.join(SCRIPT_DIR, "main.py")]
        if no_cache:
            cmd.append("--no-cache")

        env = os.environ.copy()
        env["RECONCILE_SUPPLIERS"] = ",".join(selected_suppliers)
        env["RECONCILE_BASE_DIR"] = BASE_DIR
        env["RECONCILE_MODE"] = mode
        env["PYTHONIOENCODING"] = "utf-8"

        process = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace",
            env=env, cwd=SCRIPT_DIR
        )

        for line in process.stdout:
            line = line.rstrip()
            if not line:
                continue
            run_state["logs"].append(line)

            if "处理供应商:" in line:
                supplier = line.split("处理供应商:")[-1].strip()
                run_state["current_supplier"] = supplier
                run_state["current_index"] += 1

        process.wait()

    except Exception as e:
        run_state["logs"].append(f"错误: {e}")
    finally:
        run_state["running"] = False
        run_state["done"] = True


def _parse_results_from_logs():
    """从日志中解析各供应商结果"""
    results = []
    current_supplier = ""
    for line in run_state.get("logs", []):
        if "处理供应商:" in line:
            current_supplier = line.split("处理供应商:")[-1].strip()
        if "采购对账:" in line and "一致率=" in line and current_supplier:
            purchase_rate = line.split("一致率=")[-1].strip()
            results.append({"supplier": current_supplier, "purchase": purchase_rate,
                           "finance": "", "purchase_pct": _pct(purchase_rate)})
        if "财务对账:" in line and "一致率=" in line and results:
            finance_rate = line.split("一致率=")[-1].strip()
            results[-1]["finance"] = finance_rate

    # 如果日志为空，尝试从报告文件名推断
    if not results:
        for f in _get_reports():
            if "汇总" not in f and f.endswith(".xlsx"):
                name = f.replace("对账报告_", "").split("_")[0]
                results.append({"supplier": name, "purchase": "-", "finance": "-", "purchase_pct": 0})

    return results


def _pct(s):
    try:
        return float(s.replace("%", ""))
    except Exception:
        return 0


# ═══════════════════════════════════════
# 启动
# ═══════════════════════════════════════

if __name__ == "__main__":
    ip = _get_local_ip()
    port = 5000
    print("=" * 50)
    print("  超市采购对账系统 - Web 界面")
    print("=" * 50)
    print(f"\n  请在浏览器访问:")
    print(f"    本机: http://127.0.0.1:{port}")
    print(f"    局域网: http://{ip}:{port}")
    print(f"\n  按 Ctrl+C 停止服务")
    print("=" * 50)
    app.run(host="0.0.0.0", port=port, debug=False)
