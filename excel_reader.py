# -*- coding: utf-8 -*-
"""
Excel 读取模块
读取用友系统导出的采购订单、采购入库单、进货单、退货单
"""

import re
import os
from pathlib import Path
import openpyxl


def _parse_number(val):
    """解析数字值"""
    if val is None or val == "" or val == "(空白)":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("，", "").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return None


def _read_sheet_rows(ws, start_row=1, max_empty=5):
    """读取工作表的所有行，返回列表"""
    rows = []
    empty_count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if all(v is None for v in row):
            empty_count += 1
            if empty_count >= max_empty:
                break
            continue
        empty_count = 0
        rows.append(list(row))
    return rows


def _find_header_row(rows, keywords):
    """在行列表中查找包含指定关键字的表头行，返回索引"""
    for i, row in enumerate(rows):
        row_text = " ".join(str(v) for v in row if v is not None)
        if all(kw in row_text for kw in keywords):
            return i
    return None


def _build_col_map(header_row):
    """根据表头行建立列名→索引映射"""
    col_map = {}
    for i, h in enumerate(header_row):
        if h is not None:
            key = str(h).strip().replace("\n", "").replace(" ", "")
            col_map[key] = i
    return col_map


def _get_val(row, col_map, *keys):
    """从行中按列名获取值，支持多个候选列名"""
    for key in keys:
        if key in col_map:
            idx = col_map[key]
            if idx < len(row):
                return row[idx]
    return None


# ─── 仓库名称标准化 ───
STORE_MAP = {
    "总店": "集盒超市总店仓库",
    "中职店": "集盒超市中职店仓库",
    "高职店": "集盒超市高职店仓库",
}

def _normalize_store(store_str):
    """标准化仓库名称"""
    if not store_str:
        return ""
    s = str(store_str).strip()
    # 已经是标准格式
    if "集盒超市" in s:
        return s
    # 送货单格式转换
    for short, full in STORE_MAP.items():
        if short in s:
            return full
    return s


# ═══════════════════════════════════════════════════════════
# 采购订单读取
# ═══════════════════════════════════════════════════════════
def read_purchase_order(file_path):
    """
    读取采购订单Excel
    表头第1行，列：单据日期, 单据编号, 供应商, 仓库(表头), 备注,
                   存货名称, 存货编码, 采购单位, 数量, 含税单价, 含税金额, 累计执行数量
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    items = []

    for ws in wb.worksheets:
        rows = _read_sheet_rows(ws)
        if not rows:
            continue

        header_idx = _find_header_row(rows, ["单据编号", "存货编码"])
        if header_idx is None:
            header_idx = _find_header_row(rows, ["单据编号", "存货名称"])
        if header_idx is None:
            continue

        col_map = _build_col_map(rows[header_idx])

        for row in rows[header_idx + 1:]:
            po_number = _get_val(row, col_map, "单据编号")
            barcode = _get_val(row, col_map, "存货编码")
            if not po_number or not barcode:
                continue

            items.append({
                "po_number": str(po_number).strip(),
                "date": str(_get_val(row, col_map, "单据日期") or ""),
                "supplier": str(_get_val(row, col_map, "供应商") or ""),
                "store": _normalize_store(_get_val(row, col_map, "仓库(表头)", "仓库（表头）")),
                "name": str(_get_val(row, col_map, "存货名称") or ""),
                "barcode": str(barcode).strip(),
                "unit": str(_get_val(row, col_map, "采购单位") or ""),
                "qty": _parse_number(_get_val(row, col_map, "数量")),
                "price": _parse_number(_get_val(row, col_map, "含税单价")),
                "amount": _parse_number(_get_val(row, col_map, "含税金额")),
                "executed_qty": _parse_number(_get_val(row, col_map, "累计执行数量")),
            })

    wb.close()
    return items


# ═══════════════════════════════════════════════════════════
# 采购入库单读取
# ═══════════════════════════════════════════════════════════
def read_purchase_receipt(file_path):
    """
    读取采购入库单Excel
    Sheet1（明细）列：采购订单号, 单据编号, 单据日期, 业务类型, 供应商, 备注,
                      核算状态, 仓库（表头）, 存货名称, 存货编码, 规格型号,
                      计量单位, 应收数量, 实收数量, 含税单价, 含税金额, 来源单据
    注意：翔瑞缺"核算状态"列，通力缺"应收数量"列 → 统一用"实收数量"
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    detail_items = []
    summary_items = []

    for ws in wb.worksheets:
        rows = _read_sheet_rows(ws)
        if not rows:
            continue

        # 判断是明细表还是汇总表
        first_row_text = " ".join(str(v) for v in rows[0] if v is not None) if rows else ""

        # 查找含"存货编码"或"存货名称"的表头 → 明细表
        detail_header = _find_header_row(rows, ["存货编码", "含税单价"])
        if detail_header is not None:
            col_map = _build_col_map(rows[detail_header])
            for row in rows[detail_header + 1:]:
                barcode = _get_val(row, col_map, "存货编码")
                if not barcode:
                    continue
                po = _get_val(row, col_map, "采购订单号")
                receipt_no = _get_val(row, col_map, "单据编号")

                detail_items.append({
                    "po_number": str(po).strip() if po else "",
                    "receipt_no": str(receipt_no).strip() if receipt_no else "",
                    "date": str(_get_val(row, col_map, "单据日期") or ""),
                    "biz_type": str(_get_val(row, col_map, "业务类型") or ""),
                    "supplier": str(_get_val(row, col_map, "供应商") or ""),
                    "store": _normalize_store(_get_val(row, col_map, "仓库（表头）", "仓库(表头)")),
                    "name": str(_get_val(row, col_map, "存货名称") or ""),
                    "barcode": str(barcode).strip(),
                    "spec": str(_get_val(row, col_map, "规格型号") or ""),
                    "unit": str(_get_val(row, col_map, "计量单位") or ""),
                    "expected_qty": _parse_number(_get_val(row, col_map, "应收数量")),
                    "actual_qty": _parse_number(_get_val(row, col_map, "实收数量")),
                    "price": _parse_number(_get_val(row, col_map, "含税单价")),
                    "amount": _parse_number(_get_val(row, col_map, "含税金额")),
                    "source": str(_get_val(row, col_map, "来源单据") or ""),
                })
            continue

        # 查找汇总表（含"求和项"的表头）
        summary_header = _find_header_row(rows, ["单据编号", "采购订单号"])
        if summary_header is None:
            summary_header = _find_header_row(rows, ["单据编号", "求和项"])
        if summary_header is not None:
            col_map = _build_col_map(rows[summary_header])
            for row in rows[summary_header + 1:]:
                receipt_no = _get_val(row, col_map, "单据编号")
                if not receipt_no:
                    continue
                summary_items.append({
                    "receipt_no": str(receipt_no).strip(),
                    "biz_type": str(_get_val(row, col_map, "业务类型") or ""),
                    "store": _normalize_store(_get_val(row, col_map, "仓库（表头）", "仓库(表头)")),
                    "po_number": str(_get_val(row, col_map, "采购订单号") or "").strip(),
                    "total_amount": _parse_number(_get_val(row, col_map, "求和项:含税金额", "求和项：含税金额")),
                })

    wb.close()
    return detail_items, summary_items


# ═══════════════════════════════════════════════════════════
# 进货单/退货单读取
# ═══════════════════════════════════════════════════════════
def read_goods_receipt(file_path, is_return=False):
    """
    读取进货单或退货单Excel
    表头在第8行：序号, 存货编码, 存货名称, 规格型号, 外箱装数,
                 采购单位, 数量, 含税单价, 含税金额, 采购入库单号, 采购订单号
    前面几行是单据头信息（单据日期、编号、供应商等）
    遇到"合计"行停止
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    items = []
    doc_info = {}

    for ws in wb.worksheets:
        rows = _read_sheet_rows(ws, start_row=1)
        if not rows:
            continue

        # 提取单据头信息（前7行）
        for row in rows[:8]:
            row_text = " ".join(str(v) for v in row if v is not None)
            # 单据编号
            if "单据编号" in row_text:
                for i, v in enumerate(row):
                    if v and "PS-" in str(v):
                        doc_info["doc_number"] = str(v).strip()
                        break
            # 供应商
            if "供应商" in row_text:
                for i, v in enumerate(row):
                    if v and "公司" in str(v):
                        doc_info["supplier"] = str(v).strip()
                        break
            # 单据日期
            if "单据日期" in row_text:
                for i, v in enumerate(row):
                    s = str(v) if v else ""
                    if re.match(r'\d{4}年', s):
                        doc_info["doc_date"] = s.strip()
                        break

        # 查找数据表头行（含"存货编码"和"数量"）
        header_idx = _find_header_row(rows, ["存货编码", "数量"])
        if header_idx is None:
            header_idx = _find_header_row(rows, ["序号", "存货编码"])
        if header_idx is None:
            continue

        col_map = _build_col_map(rows[header_idx])

        # 从表头下一行开始读取，遇到"合计"停止
        for row in rows[header_idx + 1:]:
            row_text = " ".join(str(v) for v in row if v is not None)
            if "合计" in row_text:
                break
            if not row_text.strip():
                continue

            barcode = _get_val(row, col_map, "存货编码")
            if not barcode:
                continue

            po = _get_val(row, col_map, "采购订单号")
            receipt_no = _get_val(row, col_map, "采购入库单号")

            items.append({
                "barcode": str(barcode).strip(),
                "name": str(_get_val(row, col_map, "存货名称") or ""),
                "spec": str(_get_val(row, col_map, "规格型号") or ""),
                "box_count": str(_get_val(row, col_map, "外箱装数") or ""),
                "unit": str(_get_val(row, col_map, "采购单位", "采购\n单位") or ""),
                "qty": _parse_number(_get_val(row, col_map, "数量")),
                "price": _parse_number(_get_val(row, col_map, "含税单价", "含税\n单价")),
                "amount": _parse_number(_get_val(row, col_map, "含税金额")),
                "receipt_no": str(receipt_no).strip() if receipt_no else "",
                "po_number": str(po).strip() if po else "",
                "is_return": is_return,
                "doc_number": doc_info.get("doc_number", ""),
                "supplier": doc_info.get("supplier", ""),
            })

    wb.close()
    return items, doc_info


# ═══════════════════════════════════════════════════════════
# 供应商文件夹自动识别
# ═══════════════════════════════════════════════════════════
def _detect_file_type_by_content(file_path):
    """
    通过读取Excel前几行内容判断文件类型
    返回: "purchase_order" / "purchase_receipt" / "goods_receipt" / "return_receipt" / None
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = []
            for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
                rows.append(list(row))

            all_text = " ".join(str(v) for row in rows for v in row if v is not None)

            # 进货单：前几行标题含"进货单"，且有单据编号 PS-
            if "进货单" in all_text and "PS-" in all_text:
                wb.close()
                return "goods_receipt"

            # 退货单：前几行标题含"退货单"
            if "退货单" in all_text and "PS-" in all_text:
                wb.close()
                return "return_receipt"

            # 采购入库单：表头含"存货编码"+"含税单价"+"采购订单号"（明细表特征）
            for row in rows:
                row_text = " ".join(str(v) for v in row if v is not None)
                if "存货编码" in row_text and "含税单价" in row_text and "采购订单号" in row_text:
                    # 区分入库单和订单：入库单有"实收数量"
                    if "实收数量" in row_text or "来源单据" in row_text:
                        wb.close()
                        return "purchase_receipt"
                    # 采购订单：有"累计执行数量"
                    if "累计执行数量" in row_text:
                        wb.close()
                        return "purchase_order"

            # 汇总表特征：含"求和项"→ 采购入库单汇总sheet
            for row in rows:
                row_text = " ".join(str(v) for v in row if v is not None)
                if "求和项" in row_text and "单据编号" in row_text:
                    wb.close()
                    return "purchase_receipt"

        wb.close()
    except Exception:
        pass
    return None


def identify_files(supplier_dir):
    """
    自动识别供应商文件夹下的各类Excel文件
    返回 dict: {
        "purchase_order": path,      # 采购订单
        "purchase_receipt": path,    # 采购入库单
        "goods_receipt": path,       # 进货单
        "return_receipt": path,      # 退货单（可能为None）
        "pdf": path,                 # PDF送货单
        "supplier_name": str,        # 供应商名称
    }
    """
    result = {
        "purchase_order": None,
        "purchase_receipt": None,
        "goods_receipt": None,
        "return_receipt": None,
        "pdf": None,
        "supplier_name": Path(supplier_dir).name,
    }

    supplier_name = Path(supplier_dir).name

    for f in Path(supplier_dir).iterdir():
        # 跳过Excel临时文件
        if f.name.startswith("~$"):
            continue
        fname = f.name.lower()

        if f.suffix.lower() == ".pdf":
            result["pdf"] = str(f)
            continue

        if f.suffix.lower() != ".xlsx":
            continue

        # ── 用友默认文件名映射（通用，不限供应商） ──
        if "puarrival" in fname:
            result["goods_receipt"] = str(f)
            continue
        if "pureceiptentry" in fname or "pureceipt" in fname:
            result["purchase_receipt"] = str(f)
            continue
        if "purchaseorder" in fname or "poorder" in fname:
            result["purchase_order"] = str(f)
            continue
        if "pureturnentry" in fname or "pureturn" in fname:
            result["return_receipt"] = str(f)
            continue

        # ── 中文关键词规则 ──
        if "退货单" in f.name:
            result["return_receipt"] = str(f)
        elif "进货单" in f.name:
            result["goods_receipt"] = str(f)
        elif "采购入库单" in f.name:
            result["purchase_receipt"] = str(f)
        elif "采购订单" in f.name:
            result["purchase_order"] = str(f)

    # ── 内容检测兜底：对未匹配的xlsx文件按内容判断类型 ──
    unmatched = []
    for f in Path(supplier_dir).iterdir():
        if f.name.startswith("~$") or f.suffix.lower() != ".xlsx":
            continue
        if str(f) not in [result["purchase_order"], result["purchase_receipt"],
                          result["goods_receipt"], result["return_receipt"]]:
            unmatched.append(f)

    for f in unmatched:
        detected = _detect_file_type_by_content(str(f))
        if detected and result[detected] is None:
            result[detected] = str(f)

    return result


def read_all_supplier_excel(supplier_dir):
    """
    读取一个供应商文件夹下的所有Excel数据
    返回 dict: {
        "purchase_order": list,      # 采购订单条目
        "purchase_receipt_detail": list,  # 采购入库单明细
        "purchase_receipt_summary": list, # 采购入库单汇总
        "goods_receipt": list,       # 进货单条目
        "return_receipt": list,      # 退货单条目
        "goods_doc_info": dict,      # 进货单单据头
        "return_doc_info": dict,     # 退货单单据头
        "files": dict,              # 文件路径映射
    }
    """
    files = identify_files(supplier_dir)
    data = {
        "purchase_order": [],
        "purchase_receipt_detail": [],
        "purchase_receipt_summary": [],
        "goods_receipt": [],
        "return_receipt": [],
        "goods_doc_info": {},
        "return_doc_info": {},
        "files": files,
    }

    if files["purchase_order"]:
        data["purchase_order"] = read_purchase_order(files["purchase_order"])

    if files["purchase_receipt"]:
        detail, summary = read_purchase_receipt(files["purchase_receipt"])
        data["purchase_receipt_detail"] = detail
        data["purchase_receipt_summary"] = summary

    if files["goods_receipt"]:
        items, doc_info = read_goods_receipt(files["goods_receipt"], is_return=False)
        data["goods_receipt"] = items
        data["goods_doc_info"] = doc_info

    if files["return_receipt"]:
        items, doc_info = read_goods_receipt(files["return_receipt"], is_return=True)
        data["return_receipt"] = items
        data["return_doc_info"] = doc_info

    return data


# ─── 测试入口 ───
if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')

    base = r"F:\claude开发项目\atutoordermatching\3.16-对账汇总"

    # 测试承运
    test_supplier = os.path.join(base, "承运")
    print("=" * 60)
    print(f"测试供应商: 承运")
    print("=" * 60)

    data = read_all_supplier_excel(test_supplier)
    files = data["files"]

    print(f"\n文件识别:")
    for k, v in files.items():
        if v:
            print(f"  {k}: {os.path.basename(v) if isinstance(v, str) else v}")

    print(f"\n采购订单: {len(data['purchase_order'])} 条")
    if data['purchase_order']:
        item = data['purchase_order'][0]
        print(f"  首条: PO={item['po_number']} 编码={item['barcode']} "
              f"{item['name'][:15]} 数量={item['qty']} 单价={item['price']} 金额={item['amount']}")
        # PO号分布
        pos = set(i['po_number'] for i in data['purchase_order'])
        print(f"  PO号: {sorted(pos)}")

    print(f"\n采购入库单明细: {len(data['purchase_receipt_detail'])} 条")
    if data['purchase_receipt_detail']:
        item = data['purchase_receipt_detail'][0]
        print(f"  首条: PO={item['po_number']} 编码={item['barcode']} "
              f"{item['name'][:15]} 实收={item['actual_qty']} 单价={item['price']} 金额={item['amount']}")
        pos = set(i['po_number'] for i in data['purchase_receipt_detail'] if i['po_number'])
        print(f"  PO号: {sorted(pos)}")

    print(f"\n采购入库单汇总: {len(data['purchase_receipt_summary'])} 条")
    for s in data['purchase_receipt_summary'][:3]:
        print(f"  {s['receipt_no']} PO={s['po_number']} 金额={s['total_amount']}")

    print(f"\n进货单: {len(data['goods_receipt'])} 条")
    if data['goods_receipt']:
        item = data['goods_receipt'][0]
        print(f"  首条: 编码={item['barcode']} {item['name'][:15]} "
              f"数量={item['qty']} 单价={item['price']} 金额={item['amount']} "
              f"入库单={item['receipt_no']} PO={item['po_number']}")
        print(f"  单据: {data['goods_doc_info']}")

    print(f"\n退货单: {len(data['return_receipt'])} 条")

    # 测试太古（特殊文件名）
    print("\n" + "=" * 60)
    print("测试供应商: 太古")
    print("=" * 60)
    test_supplier2 = os.path.join(base, "太古")
    data2 = read_all_supplier_excel(test_supplier2)
    files2 = data2["files"]
    print(f"\n文件识别:")
    for k, v in files2.items():
        if v:
            print(f"  {k}: {os.path.basename(v) if isinstance(v, str) else v}")
    print(f"\n采购订单: {len(data2['purchase_order'])} 条")
    print(f"采购入库单明细: {len(data2['purchase_receipt_detail'])} 条")
    print(f"进货单: {len(data2['goods_receipt'])} 条")

    # 测试汇基（有退货单）
    print("\n" + "=" * 60)
    print("测试供应商: 汇基（有退货单）")
    print("=" * 60)
    test_supplier3 = os.path.join(base, "汇基")
    data3 = read_all_supplier_excel(test_supplier3)
    print(f"\n进货单: {len(data3['goods_receipt'])} 条")
    print(f"退货单: {len(data3['return_receipt'])} 条")
    if data3['return_receipt']:
        item = data3['return_receipt'][0]
        print(f"  首条退货: 编码={item['barcode']} {item['name'][:15]} "
              f"数量={item['qty']} 金额={item['amount']}")

    # 测试所有供应商的文件识别
    print("\n" + "=" * 60)
    print("所有供应商文件识别汇总")
    print("=" * 60)
    for d in sorted(Path(base).iterdir()):
        if d.is_dir():
            files = identify_files(str(d))
            missing = [k for k in ["purchase_order", "purchase_receipt", "goods_receipt", "pdf"]
                       if not files[k]]
            ret = "有退货单" if files["return_receipt"] else ""
            miss = f" 缺少: {missing}" if missing else ""
            print(f"  {d.name}: {ret}{miss}" if ret or miss else f"  {d.name}: 完整")
