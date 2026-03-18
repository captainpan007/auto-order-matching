# -*- coding: utf-8 -*-
"""
对账报告生成模块
将比对结果输出为格式化的Excel报告
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── 样式定义 ───
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_BLUE = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_BOLD = Font(bold=True, size=11)
FONT_NORMAL = Font(size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _write_header(ws, row, headers):
    """写入表头行"""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def _write_row(ws, row, values, fill=None):
    """写入数据行"""
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER_THIN
        if fill:
            cell.fill = fill


def _auto_width(ws, min_width=8, max_width=30):
    """自动调整列宽"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # 中文字符算2个宽度
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = max(min(max_len + 2, max_width), min_width)


def generate_supplier_report(result, output_dir, date_str=None):
    """
    生成单个供应商的对账报告

    Args:
        result: reconcile_supplier()的返回值
        output_dir: 输出目录
        date_str: 日期字符串（默认今天）

    Returns:
        str: 生成的报告文件路径
    """
    if not date_str:
        date_str = datetime.now().strftime("%Y%m%d")

    supplier = result["supplier_name"]
    p1 = result["phase1"]
    p2 = result["phase2"]

    wb = Workbook()

    # ═══ Sheet1: 一致项 ═══
    ws1 = wb.active
    ws1.title = "一致项"
    headers = [
        "序号", "对账阶段", "PO号", "存货编码", "商品名称(我方)", "商品名称(送货单)",
        "仓库", "数量(送货单)", "数量(我方)", "单价(送货单)", "单价(我方)",
        "金额(送货单)", "金额(我方)", "结果"
    ]
    _write_header(ws1, 1, headers)

    row = 2
    # 阶段一一致项
    for item in p1["matched"]:
        _write_row(ws1, row, [
            row - 1, "采购对账", item["po_number"], item["barcode"],
            item["name_receipt"], item["name_delivery"], item["store"],
            item["qty_delivery"], item["qty_receipt"],
            item["price_delivery"], item["price_receipt"],
            item["amount_delivery"], item["amount_receipt"], "一致"
        ], fill=FILL_GREEN)
        row += 1

    # 阶段二一致项（去重：与阶段一相同的条目不重复列出）
    p1_keys = set((i["po_number"], i["barcode"]) for i in p1["matched"])
    for item in p2["matched"]:
        key = (item["po_number"], item["barcode"])
        if key in p1_keys:
            continue
        _write_row(ws1, row, [
            row - 1, "财务对账", item["po_number"], item["barcode"],
            item["name_goods"], item["name_delivery"], "",
            item["qty_delivery"], item["qty_goods"],
            item["price_delivery"], item["price_goods"],
            item["amount_delivery"], item["amount_goods"], "一致"
        ], fill=FILL_GREEN)
        row += 1

    _auto_width(ws1)

    # ═══ Sheet2: 差异项 ═══
    ws2 = wb.create_sheet("差异项")
    headers2 = [
        "序号", "对账阶段", "严重级别", "PO号", "存货编码",
        "商品名称(我方)", "商品名称(送货单)", "仓库",
        "差异字段", "送货单值", "我方值", "差值",
        "数量(送货单)", "数量(我方)", "单价(送货单)", "单价(我方)",
        "金额(送货单)", "金额(我方)", "OCR备注"
    ]
    _write_header(ws2, 1, headers2)

    row = 2
    all_diffs = []
    for item in p1["diff"]:
        all_diffs.append(("采购对账", item, item.get("name_receipt",""), item.get("name_delivery","")))
    for item in p2["diff"]:
        all_diffs.append(("财务对账", item, item.get("name_goods",""), item.get("name_delivery","")))

    for phase, item, name_ours, name_delivery in all_diffs:
        fill = FILL_RED if item.get("severity") == "红色" else FILL_YELLOW

        # 构建差异详情
        diff_detail = ""
        our_val = ""
        their_val = ""

        if "数量" in item.get("diff_fields", ""):
            qty_key_ours = "qty_receipt" if phase == "采购对账" else "qty_goods"
            diff_detail = "数量"
            their_val = item.get("qty_delivery")
            our_val = item.get(qty_key_ours)
        if "单价" in item.get("diff_fields", ""):
            price_key_ours = "price_receipt" if phase == "采购对账" else "price_goods"
            diff_detail = ("单价" if not diff_detail else diff_detail + ", 单价")
            their_val = item.get("price_delivery")
            our_val = item.get(price_key_ours)
        if "金额" in item.get("diff_fields", ""):
            amount_key_ours = "amount_receipt" if phase == "采购对账" else "amount_goods"
            diff_detail = ("金额" if not diff_detail else diff_detail + ", 金额")
            their_val = item.get("amount_delivery")
            our_val = item.get(amount_key_ours)

        qty_key_ours = "qty_receipt" if phase == "采购对账" else "qty_goods"
        price_key_ours = "price_receipt" if phase == "采购对账" else "price_goods"
        amount_key_ours = "amount_receipt" if phase == "采购对账" else "amount_goods"

        _write_row(ws2, row, [
            row - 1, phase, item.get("severity", ""),
            item["po_number"], item["barcode"],
            name_ours, name_delivery, item.get("store", ""),
            item.get("diff_fields", ""),
            their_val, our_val,
            item.get("qty_diff") or item.get("price_diff") or item.get("amount_diff"),
            item.get("qty_delivery"), item.get(qty_key_ours),
            item.get("price_delivery"), item.get(price_key_ours),
            item.get("amount_delivery"), item.get(amount_key_ours),
            item.get("ocr_warning", ""),
        ], fill=fill)
        row += 1

    _auto_width(ws2)

    # ═══ Sheet3: 未匹配项 ═══
    ws3 = wb.create_sheet("未匹配项")
    headers3 = [
        "序号", "来源", "对账阶段", "PO号", "存货编码", "商品名称",
        "仓库", "数量", "单价", "金额", "可能原因", "OCR备注"
    ]
    _write_header(ws3, 1, headers3)

    row = 2
    all_unmatched = []
    for item in p1["unmatched_delivery"]:
        all_unmatched.append(("采购对账", item))
    for item in p1["unmatched_receipt"]:
        all_unmatched.append(("采购对账", item))
    for item in p2["unmatched_goods"]:
        all_unmatched.append(("财务对账", item))
    for item in p2["unmatched_delivery"]:
        all_unmatched.append(("财务对账", item))

    for phase, item in all_unmatched:
        fill = FILL_GRAY if "OCR" in item.get("ocr_warning", "") else FILL_RED
        _write_row(ws3, row, [
            row - 1, item.get("source", ""), phase,
            item.get("po_number", ""), item.get("barcode", ""),
            item.get("name", ""), item.get("store", ""),
            item.get("qty"), item.get("price"), item.get("amount"),
            item.get("reason", ""), item.get("ocr_warning", ""),
        ], fill=fill)
        row += 1

    _auto_width(ws3)

    # ═══ Sheet4: 退货明细 ═══
    ws4 = wb.create_sheet("退货明细")
    headers4 = [
        "序号", "PO号", "存货编码", "商品名称", "入库单号",
        "退货数量", "退货单价", "退货金额"
    ]
    _write_header(ws4, 1, headers4)

    row = 2
    for item in p2.get("returns", []):
        _write_row(ws4, row, [
            row - 1, item.get("po_number", ""), item.get("barcode", ""),
            item.get("name", ""), item.get("receipt_no", ""),
            item.get("qty"), item.get("price"), item.get("amount"),
        ], fill=FILL_BLUE)
        row += 1

    if row == 2:
        ws4.cell(row=2, column=1, value="无退货记录").font = FONT_NORMAL

    _auto_width(ws4)

    # ═══ Sheet5: 对账汇总 ═══
    ws5 = wb.create_sheet("对账汇总")

    ws5.cell(row=1, column=1, value="对账汇总报告").font = Font(bold=True, size=14)
    ws5.merge_cells("A1:D1")

    info = [
        ("供应商", supplier),
        ("对账日期", date_str),
        ("", ""),
        ("【阶段一：采购对账】", "送货单 vs 采购入库单"),
        ("送货单条目数", p1["summary"]["total_delivery"]),
        ("入库单条目数", p1["summary"]["total_receipt"]),
        ("一致条目", p1["summary"]["matched"]),
        ("差异条目", p1["summary"]["diff"]),
        ("送货单未匹配", p1["summary"]["unmatched_delivery"]),
        ("入库单未匹配", p1["summary"]["unmatched_receipt"]),
        ("一致率", p1["summary"]["match_rate"]),
        ("", ""),
        ("【阶段二：财务对账】", "进货单 vs 送货单 vs 退货单"),
        ("进货单条目数", p2["summary"]["total_goods"]),
        ("送货单条目数", p2["summary"]["total_delivery"]),
        ("退货单条目数", p2["summary"]["total_returns"]),
        ("一致条目", p2["summary"]["matched"]),
        ("差异条目", p2["summary"]["diff"]),
        ("进货单未匹配", p2["summary"]["unmatched_goods"]),
        ("送货单未匹配", p2["summary"]["unmatched_delivery"]),
        ("一致率", p2["summary"]["match_rate"]),
        ("", ""),
        ("【金额校验】", ""),
        ("进货单总额", p2["amount_check"]["goods_total"]),
        ("送货单总额", p2["amount_check"]["delivery_total"]),
        ("退货单总额", p2["amount_check"]["return_total"]),
        ("实际应付金额", p2["amount_check"]["actual_payable"]),
        ("文件名预期金额", p2["amount_check"]["file_amount"]),
        ("进货-送货差额", p2["amount_check"]["goods_vs_delivery_diff"]),
        ("应付-预期差额", p2["amount_check"]["payable_vs_file_diff"]),
    ]

    for i, (label, value) in enumerate(info, 3):
        cell_a = ws5.cell(row=i, column=1, value=label)
        cell_b = ws5.cell(row=i, column=2, value=value)
        if label.startswith("【"):
            cell_a.font = FONT_BOLD
            cell_b.font = FONT_BOLD
        else:
            cell_a.font = FONT_NORMAL
            cell_b.font = FONT_NORMAL

    _auto_width(ws5)

    # 保存
    os.makedirs(output_dir, exist_ok=True)
    filename = f"对账报告_{supplier}_{date_str}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


def generate_summary_report(all_results, output_dir, date_str=None):
    """
    生成所有供应商的汇总报告

    Args:
        all_results: 所有供应商的对账结果列表
        output_dir: 输出目录
        date_str: 日期字符串

    Returns:
        str: 生成的报告文件路径
    """
    if not date_str:
        date_str = datetime.now().strftime("%Y%m%d")

    wb = Workbook()
    ws = wb.active
    ws.title = "对账汇总"

    headers = [
        "供应商", "对账日期",
        "送货单条目", "入库单条目", "进货单条目", "退货单条目",
        "采购对账-一致", "采购对账-差异", "采购对账-未匹配", "采购对账-一致率",
        "财务对账-一致", "财务对账-差异", "财务对账-未匹配", "财务对账-一致率",
        "进货单总额", "送货单总额", "退货总额", "实际应付",
        "文件名金额", "应付vs文件差额",
    ]
    _write_header(ws, 1, headers)

    row = 2
    totals = {k: 0 for k in [
        "d", "r", "g", "ret",
        "p1m", "p1d", "p1u",
        "p2m", "p2d", "p2u",
        "gt", "dt", "rtt", "ap",
    ]}

    for result in all_results:
        p1 = result["phase1"]["summary"]
        p2 = result["phase2"]["summary"]
        ac = result["phase2"]["amount_check"]

        p1_unmatched = p1["unmatched_delivery"] + p1["unmatched_receipt"]
        p2_unmatched = p2["unmatched_goods"] + p2["unmatched_delivery"]

        values = [
            result["supplier_name"], date_str,
            p1["total_delivery"], p1["total_receipt"],
            p2["total_goods"], p2["total_returns"],
            p1["matched"], p1["diff"], p1_unmatched, p1["match_rate"],
            p2["matched"], p2["diff"], p2_unmatched, p2["match_rate"],
            ac["goods_total"], ac["delivery_total"],
            ac["return_total"], ac["actual_payable"],
            ac["file_amount"], ac["payable_vs_file_diff"],
        ]
        _write_row(ws, row, values)
        row += 1

        totals["d"] += p1["total_delivery"]
        totals["r"] += p1["total_receipt"]
        totals["g"] += p2["total_goods"]
        totals["ret"] += p2["total_returns"]
        totals["p1m"] += p1["matched"]
        totals["p1d"] += p1["diff"]
        totals["p1u"] += p1_unmatched
        totals["p2m"] += p2["matched"]
        totals["p2d"] += p2["diff"]
        totals["p2u"] += p2_unmatched
        totals["gt"] += ac["goods_total"]
        totals["dt"] += ac["delivery_total"]
        totals["rtt"] += ac["return_total"]
        totals["ap"] += ac["actual_payable"]

    # 合计行
    p1_rate = f"{totals['p1m']/max(totals['d'],1)*100:.1f}%"
    p2_rate = f"{totals['p2m']/max(totals['g'],1)*100:.1f}%"
    total_row = [
        "合计", date_str,
        totals["d"], totals["r"], totals["g"], totals["ret"],
        totals["p1m"], totals["p1d"], totals["p1u"], p1_rate,
        totals["p2m"], totals["p2d"], totals["p2u"], p2_rate,
        round(totals["gt"], 2), round(totals["dt"], 2),
        round(totals["rtt"], 2), round(totals["ap"], 2),
        "", "",
    ]
    _write_row(ws, row, total_row, fill=FILL_YELLOW)
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).font = FONT_BOLD

    _auto_width(ws)

    os.makedirs(output_dir, exist_ok=True)
    filename = f"对账汇总报告_{date_str}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ─── 测试入口 ───
if __name__ == "__main__":
    import sys, json
    sys.stdout.reconfigure(encoding='utf-8')

    from excel_reader import read_all_supplier_excel
    from reconciler import reconcile_supplier

    base = r"F:\claude开发项目\atutoordermatching\3.16-对账汇总"
    output_dir = r"F:\claude开发项目\atutoordermatching\对账结果"
    supplier = "承运"
    supplier_dir = os.path.join(base, supplier)

    # 读取数据
    excel_data = read_all_supplier_excel(supplier_dir)
    cache_file = os.path.join(supplier_dir, "_ocr_cache.json")
    with open(cache_file, 'r', encoding='utf-8') as f:
        delivery_items = json.load(f)

    # 比对
    result = reconcile_supplier(delivery_items, excel_data, supplier)

    # 生成报告
    report_path = generate_supplier_report(result, output_dir)
    print(f"供应商报告已生成: {report_path}")

    # 生成汇总报告（仅承运）
    summary_path = generate_summary_report([result], output_dir)
    print(f"汇总报告已生成: {summary_path}")
